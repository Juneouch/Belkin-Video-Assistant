"""Build a director workbook from structured JSON.

Usage:
  python build_director_workbook.py input.json output.xlsx

The JSON may contain: project, shotlist, characters, locations, props, product,
prompts, storyboard_index. Each collection is a list of dictionaries. Missing
collections are created with the standard headers so the workbook remains editable.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


SHEETS = {
    "Shotlist": [
        "Shot ID", "Scene ID", "Beat", "Type", "Time In", "Time Out", "Duration",
        "Aspect / Crop", "Shot Size", "Angle / Height", "Lens / FOV", "Camera Movement",
        "Composition", "Action / Performance", "Product Role", "Feature / Proof",
        "Character IDs", "Location ID", "Prop IDs", "Lighting", "Color / Mood",
        "Material / Surface", "Audio / SFX / VO", "On-screen Copy", "Edit / Transition",
        "Prompt ID", "Reference IDs", "Status", "Notes",
    ],
    "Characters": [
        "Character ID", "Name / Role", "Narrative Function", "Age Range", "Appearance",
        "Wardrobe", "Silhouette / Scale", "Gesture Language", "Performance Arc",
        "Relationship to Product", "Continuity Constraints", "Reference IDs", "Status", "Notes",
    ],
    "Locations": [
        "Location ID", "Name", "Narrative Function", "Spatial Description", "Time / Weather",
        "Architecture", "Palette", "Key Surfaces", "Practical / Ambient Elements",
        "Light Direction", "Camera Access", "Motion / Atmosphere", "Continuity Constraints",
        "Reference IDs", "Status", "Notes",
    ],
    "Props": [
        "Prop ID", "Name", "Narrative Function", "Shot IDs", "Interaction", "State Before",
        "State After", "Color / Material", "Scale / Placement", "Continuity Constraints",
        "Competes with Product?", "Reference IDs", "Status", "Notes",
    ],
    "Product": [
        "Product ID / SKU", "Variant", "Approved Name", "Hero Feature", "Feature Proof",
        "Silhouette / Hero Orientation", "Color / Finish", "Material Truth",
        "Ports / Buttons / UI Truth", "Reference Image IDs", "Do Not Change", "Claims / Legal Notes",
        "Status", "Notes",
    ],
    "Prompt Pack": [
        "Prompt ID", "Shot ID", "Frame Purpose", "Product Lock", "Character Lock", "Location Lock",
        "Prop Lock", "Action", "Composition", "Camera", "Lighting", "Material", "Mood / Valence",
        "Arousal", "Color", "Continuity", "Negative Prompt", "Reference IDs", "Image Path", "Approval", "Notes",
    ],
    "Storyboard Index": [
        "Frame ID", "Prompt ID", "Shot ID", "Scene ID", "Image Path / Link", "Aspect",
        "Generation Version", "Source References", "Approval", "Revision Note", "Owner", "Date",
    ],
}

HEADER_FILL = PatternFill("solid", fgColor="102A43")
ACCENT_FILL = PatternFill("solid", fgColor="D9EAF7")
WHITE_FONT = Font(color="FFFFFF", bold=True)
THIN_GRAY = Side(style="thin", color="D9E2EC")


def value_for(row: dict[str, Any], header: str) -> Any:
    if header in row:
        return row[header]
    normalized = header.lower().replace("/", " ").replace("?", "").replace(" ", "_")
    for key, value in row.items():
        candidate = str(key).lower().replace("/", " ").replace("?", "").replace(" ", "_")
        if candidate == normalized:
            return value
    return ""


def write_table(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for header_cell in ws[1]:
        header_cell.fill = HEADER_FILL
        header_cell.font = WHITE_FONT
        header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_cell.border = Border(bottom=THIN_GRAY)
    for row in rows:
        ws.append([value_for(row, header) for header in headers])
    if ws.max_row >= 2:
        ref = f"A1:{chr(64 + min(ws.max_column, 26))}{ws.max_row}"
        if ws.max_column > 26:
            from openpyxl.utils import get_column_letter
            ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        table = Table(displayName=f"Table_{ws.title.replace(' ', '_')}", ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in ws.columns:
        letter = column[0].column_letter
        max_len = max(len(str(cell.value or "")) for cell in column[: min(ws.max_row, 25)])
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 34)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_workbook(data: dict[str, Any], output: Path) -> None:
    wb = Workbook()
    readme = wb.active
    readme.title = "Read Me"
    readme.append(["Belkin Video Assistant Workbook"])
    readme["A1"].font = Font(size=16, bold=True, color="102A43")
    readme["A3"] = "Project"
    readme["B3"] = data.get("project", {}).get("name", "Untitled project")
    readme["A4"] = "Version"
    readme["B4"] = data.get("project", {}).get("version", "v1")
    readme["A5"] = "Runtime"
    readme["B5"] = data.get("project", {}).get("runtime", "TBD")
    readme["A6"] = "Source treatment"
    readme["B6"] = data.get("project", {}).get("source", "TBD")
    readme["A8"] = "Assumptions / open approvals"
    readme["A8"].fill = ACCENT_FILL
    readme["A8"].font = Font(bold=True, color="102A43")
    assumptions = data.get("project", {}).get("assumptions", [])
    for item in assumptions or ["Add director assumptions and open approvals here."]:
        readme.append([str(item)])
    readme.column_dimensions["A"].width = 34
    readme.column_dimensions["B"].width = 80
    for row in readme.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for name, headers in SHEETS.items():
        ws = wb.create_sheet(name)
        rows = data.get(name.lower().replace(" ", "_"), data.get(name, [])) or []
        write_table(ws, headers, rows)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input_json", type=Path)
    parser.add_argument("output_xlsx", type=Path)
    args = parser.parse_args()
    data = json.loads(args.input_json.read_text(encoding="utf-8"))
    build_workbook(data, args.output_xlsx)
    print(f"Created {args.output_xlsx}")


if __name__ == "__main__":
    main()
